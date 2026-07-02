"""
Report Generator Agent – produces the final Excel compliance report.
"""

from __future__ import annotations

import logging
from collections import Counter
from typing import TYPE_CHECKING

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config

if TYPE_CHECKING:
    from agents.log_follower import LogFollowerAgent
    from agents.orchestrator import AnchoredViolation

logger = logging.getLogger(__name__)

# Severity colour map
SEVERITY_COLOURS = {
    "High":         "FFCCCC",  # light red
    "Medium":       "FFF3CC",  # light yellow
    "Low":          "CCFFCC",  # light green
    "Needs Review": "E6E6FA",  # lavender (for uncertain cases)
}


class ReportGeneratorAgent:
    """Generates compliance_report.xlsx from AnchoredViolation records."""

    def __init__(self, log_agent: "LogFollowerAgent") -> None:
        self.log = log_agent

    def generate(self, violations: list["AnchoredViolation"], total_rules_checked: int) -> None:
        self.log.log(
            "ReportGeneratorAgent",
            "generate_start",
            input_summary=f"{len(violations)} violations, {total_rules_checked} checks",
        )

        # ── Build main DataFrame ────────────────────────────────────────────
        rows = []
        for v in violations:
            rows.append({
                "File Name":        v.file_name,
                "Function":         v.function,
                "Line Number":      v.line,
                "Column":           v.column,
                "Rule ID":          v.rule_id,
                "Rule Description": v.rule_description,
                "Category":         v.rule_category,
                "Severity":         v.severity,
                "Violation":        "YES" if v.violation else "NO",
                "Confidence":       f"{v.confidence:.2f}",
                "Detection Source": v.detection_source,
                "Suggested Fix":    v.suggested_fix,
                "Rationale":        v.rationale,
            })

        main_df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
            "File Name", "Function", "Line Number", "Column",
            "Rule ID", "Rule Description", "Category", "Severity",
            "Violation", "Confidence", "Detection Source",
            "Suggested Fix", "Rationale",
        ])

        # ── Summary stats ────────────────────────────────────────────────────
        confirmed = [v for v in violations if v.violation]
        severity_dist = Counter(v.severity for v in confirmed)
        compliance_score = round(
            (1 - len(confirmed) / max(total_rules_checked, 1)) * 100, 1
        )

        summary_data = {
            "Metric": [
                "Total Rule Checks",
                "Violations Found",
                "High Severity",
                "Medium Severity",
                "Low Severity",
                "Compliance Score (%)",
            ],
            "Value": [
                total_rules_checked,
                len(confirmed),
                severity_dist.get("High", 0),
                severity_dist.get("Medium", 0),
                severity_dist.get("Low", 0),
                compliance_score,
            ],
        }
        summary_df = pd.DataFrame(summary_data)

        # ── Write to Excel ────────────────────────────────────────────────────
        config.OUTPUT_DIR.mkdir(exist_ok=True)
        with pd.ExcelWriter(config.REPORT_PATH, engine="openpyxl") as writer:
            main_df.to_excel(writer, sheet_name="Violations", index=False)
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            self._format_violations(writer.sheets["Violations"], main_df)
            self._format_summary(writer.sheets["Summary"])

        logger.info("Report saved: %s", config.REPORT_PATH)
        self.log.log(
            "ReportGeneratorAgent",
            "generate_complete",
            output_summary=f"Report: {config.REPORT_PATH}",
        )

    # ── Formatting helpers ────────────────────────────────────────────────────

    def _format_violations(self, ws, df: pd.DataFrame) -> None:
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = border

        sev_col = None
        viol_col = None
        for col_idx, col_name in enumerate(df.columns, 1):
            if col_name == "Severity":
                sev_col = col_idx
            if col_name == "Violation":
                viol_col = col_idx

        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = border

            if sev_col:
                sev_val = ws.cell(row=row_idx, column=sev_col).value or ""
                colour = SEVERITY_COLOURS.get(sev_val, "FFFFFF")
                ws.cell(row=row_idx, column=sev_col).fill = PatternFill("solid", fgColor=colour)

            if viol_col:
                viol_val = ws.cell(row=row_idx, column=viol_col).value or ""
                if viol_val == "YES":
                    ws.cell(row=row_idx, column=viol_col).fill = PatternFill("solid", fgColor="FFB3B3")

        # Auto-width columns (capped)
        col_widths = {"Rationale": 60, "Suggested Fix": 50, "Rule Description": 50}
        for col_idx, col_name in enumerate(df.columns, 1):
            letter = get_column_letter(col_idx)
            width = col_widths.get(col_name, 20)
            ws.column_dimensions[letter].width = width

        ws.freeze_panes = "A2"

    def _format_summary(self, ws) -> None:
        header_fill = PatternFill("solid", fgColor="2E7D32")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 30
